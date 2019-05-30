import openpyxl as op
import datetime

days = ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']

filename = 'work-sample.xlsx'
workbook = op.load_workbook(filename)
sheet1 = workbook.active

def shifts(row, num):
  col = 0
  for i in range(num):
    sheet1.cell(row=row, column=(3+(num*i))).value = 'Start'
    sheet1.cell(row=row, column=(4+(num*i))).value = 'End'
    if i == num-1:
      col = 5 + (num * i)
  sheet1.cell(row=row, column=col).value = 'Hours'
def week(row):
  for i in range(len(days)):
    sheet1.cell(row=row+i, column=1).value = days[i]
def dates(row, col):
  start_date = sheet1.cell(row=row, column=col).value
  for i in range(1, len(days)):
    start_date += datetime.timedelta(days=1)
    sheet1.cell(row=row+i, column=col).value = start_date
def init(r, c):
  sheet1.cell(row=r, column=1).value = 'Day'
  sheet1.cell(row=r, column=2).value = 'Date'
  shifts(r, 2)
  week(r+1)
  sheet1.cell(row=r+1, column=2).value = datetime.date(datetime.date.today().year, 5, 26)
  dates(r+1, 2)
def main():
  i = 1
  while sheet1.cell(row=i, column=1).value is not None:
    i += 1
  i += 1
  init(i, 1)

if sheet1.cell(row=1, column=1).value is None:
  init(1, 1)
else:
  main()
workbook.close

for i in range(1, 10):
  for j in range(1, 10):
    print(i, j, sheet1.cell(row=i, column=j).value)
